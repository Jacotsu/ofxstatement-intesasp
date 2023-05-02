#!/usr/bin/env python3.6

import logging
import os
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal

from ofxstatement.parser import StatementParser
from ofxstatement.plugin import Plugin
from ofxstatement.statement import (Statement, StatementLine,
                                    generate_transaction_id)
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger('IntesaSP')

GITHUB_URL = "https://github.com/Jacotsu/ofxstatement-intesasp/issues"

"""
Data wrapper general
"""


class Movimento:
    stat_line: StatementLine = None


"""
Data wrapper for Version 1 (Movimenti_Conto_<DATE>.xlsx)
"""


@dataclass
class Movimento_V1(Movimento):
    data_contabile: datetime
    data_valuta: datetime
    descrizione: str
    accrediti: Decimal
    addebiti: Decimal
    descrizione_estesa: str
    mezzo: str

    def __post_init__(self):
        # Modificare descrizione_estesa per comprendere sempre la descrizione
        self.descrizione_estesa = f"({self.descrizione}) " \
                                  f"{self.descrizione_estesa}"

        # Una volta raccolti i dati, li formatto nello standard corretto
        self.stat_line = StatementLine(
            None,
            self.data_contabile,
            self.descrizione_estesa,
            Decimal(self.accrediti) if self.accrediti else Decimal(self.addebiti)
        )
        self.stat_line.id = generate_transaction_id(self.stat_line)
        self.stat_line.date_user = self.data_valuta
        self.stat_line.trntype = self._get_transaction_type()

    def _get_transaction_type(self):
        # OFX Spec https://financialdataexchange.org/ofx
        # 11.4.4.3 Transaction Types Used in <TRNTYPE>
        # !!! Write All key value lower case !!!
        # Map value present in ~/.local/lib/python3.10/site-packages/ofxstatement/statement.py

        trans_map = {
            # Pagamenti POS
            'pagamento pos': 'POS',
            'pagamento tramite pos': 'POS',
            'pagamento effettuato su pos estero': 'POS',
            'storno pagamento pos': 'POS',
            'storno pagamento pos estero': 'POS',
            # Accrediti
            'canone mensile base e servizi aggiuntivi': 'SRVCHG',
            'prelievo carta debito su banche del gruppo': 'CASH',
            'prelievo carta debito su banche italia/sepa': 'CASH',
            'comm.prelievo carta debito italia/sepa': 'SRVCHG',
            'stipendio o pensione': 'CREDIT',
            # Ricariche Cellulari
            'pagamento mav via internet banking': 'PAYMENT',
            'pagamento bolletta cbill': 'PAYMENT',
            'pagamento telefono': 'PAYMENT',
            'ricarica tramite internet:vodafonecard': 'PAYMENT',
            'ricarica tramite internet:windtre': 'PAYMENT',
            # Commissioni Pagamento
            'commissione bolletta cbill': 'FEE',
            'commissioni bollettino postale via internet': 'FEE',
            'commissioni e spese adue': 'FEE',
            'commissioni su pagamento via internet': 'FEE',
            'imposta di bollo e/c e rendiconto': 'FEE',
            'commiss. su beu internet banking': 'FEE',
            # Operazioni Bancarie
            'accredito beu con contabile': 'XFER',
            'beu tramite internet banking': 'XFER',
            'versamento contanti su sportello automatico': 'ATM',
            'canone annuo o-key sms': 'SRVCHG',
            'pagamento adue': 'DIRECTDEBIT',
            'rata bonif. periodico con contab.': 'REPEATPMT',
            'bonifico in euro verso ue/sepa canale telem.': 'PAYMENT',
            'accredito bonifico istantaneo': 'DIRECTDEP',
            'pagamenti disposti su circuito fast pay': 'PAYMENT',
            'pagamento bollettino postale via internet': 'PAYMENT',
            'pagamento via internet': 'DIRECTDEBIT',
            # Other
            'donazione preautorizzata ad ente no profit': 'DIRECTDEBIT',
            'add. deleghe fisco/inps/regioni': 'DEBIT',
            'pagamento delega f24 via internet banking': 'PAYMENT',
        }
        try:
            cur_transaction = trans_map[self.descrizione.lower()]
        except KeyError:
            cur_transaction = 'DIRECTDEBIT'
            logging.warning(
                f"Unknown transaction type: '{self.descrizione}', "
                f"assigning default transaction type '{cur_transaction}'\n"
                f"PLEASE open an issue on GitHub '{GITHUB_URL}' "
                "in order to help us fix it"
            )
        return cur_transaction


@dataclass
class Movimento_V2(Movimento):
    data: datetime
    operazione: str
    dettagli: str
    conto_carta: str
    contabilizzazione: str
    categoria: str
    valuta: str
    importo: Decimal

    def __post_init__(self):
        # Modificare descrizione_estesa per comprendere sempre la categoria
        descrizione_estesa = f"[({self.categoria})-({self.operazione})] " \
                             f"{self.dettagli}"

        # Una volta raccolti i dati, li formatto nello standard corretto
        # NB: Questo formato non prevede la differenziazione tra la data
        #     contabile e quella di valuta, tutte le transazione useranno la
        #     data di valuta come riferimento
        self.stat_line = StatementLine(
            None, self.data, descrizione_estesa, Decimal(self.importo)
        )
        self.stat_line.id = generate_transaction_id(self.stat_line)
        self.stat_line.date_user = self.data
        self.stat_line.trntype = self._get_transaction_type()

    def _get_transaction_type(self):
        # OFX Spec https://financialdataexchange.org/ofx
        # 11.4.4.3 Transaction Types Used in <TRNTYPE>
        # !!! Write All key value lower case !!!
        # Map value present in ~/.local/lib/python3.10/site-packages/ofxstatement/statement.py

        category_map = {
            'addebiti vari': 'DEBIT',
            'entrate varie': 'CREDIT',
            'abbigliamento e accessori': 'POS',
            'altre uscite': 'POS',
            'associazioni': 'DIRECTDEBIT',
            'bonifici in uscita': 'XFER',
            'bonifici ricevuti': 'XFER',
            'carburanti': 'PAYMENT',
            'casa varie': 'POS',
            'cellulare': 'REPEATPMT',
            'cliniche': 'POS',
            'corsi e sport': 'POS',
            'cura della persona': 'POS',
            'domiciliazioni e utenze': 'REPEATPMT',
            'donazioni': 'DIRECTDEBIT',
            'farmacia': 'POS',
            'generi alimentari e supermercato': 'POS',
            'hi-tech e informatica': 'POS',
            'imposte sul reddito e tasse varie': 'FEE',
            'tasse varie': 'FEE',
            'commissioni': 'SRVCHG',
            'imposte, bolli e commissioni': 'FEE',
            'pedaggi e telepass': 'FEE',
            'rate mutuo e finanziamento': 'REPEATPMT',
            'regali ricevuti': 'CREDIT',
            'rimborsi spese e storni': 'CREDIT',
            'ristoranti e bar': 'POS',
            'spese mediche': 'POS',
            'spettacoli e musei': 'POS',
            'stipendi e pensioni': 'DIRECTDEP',
            'tv, internet, telefono': 'POS',
            'tabaccai e simili': 'POS',
            'tempo libero varie': 'POS',
            'viaggi e vacanze': 'POS'
        }

        try:
            cur_transaction = category_map[self.categoria.lower()]
        except KeyError:
            cur_transaction = 'CREDIT' if self.importo >= 0 else 'DEBIT'
            logging.warning(
                f"Unknown category: '{self.categoria}', "
                f"assigning generic category: '{cur_transaction}'"
                f"PLEASE open an issue on GitHub '{GITHUB_URL}' "
                "in order to help us fix it"
            )
        return cur_transaction


class IntesaSanPaoloPlugin(Plugin):

    def get_parser(self, filename):
        parser = IntesaSanPaoloXlsxParser(filename, self.settings)
        return parser


class IntesaSanPaoloXlsxParser(StatementParser):
    excel_version: int = None
    wb = None
    settings = None

    def __init__(self, filename, settings):
        self.file = filename
        self.wb = load_workbook(self.file)
        self.settings = settings
        logging.debug(self.settings)

        if 'Lista Movimenti' in self.wb.sheetnames:
            logging.debug('Detected "Lista Movimenti", using V1 excel parser')
            self.excel_version = 1
        elif 'Lista Operazione' in self.wb.sheetnames:
            logging.debug('Detected "Lista Operazione", using V2 excel parser')
            self.excel_version = 2
        else:
            logging.error('Unknown excel format, aborting')
            exit(os.EX_IOERR)

        self.statement = Statement(
            settings.get('abi'),
            self._get_account_id(),
            self._get_currency()
        )
        self.statement.start_balance = self._get_start_balance()
        self.statement.start_date = self._get_start_date()
        self.statement.end_balance = self._get_end_balance()
        self.statement.end_date = self._get_end_date()
        logging.debug(self.statement)

    """
    Override method, use to obrain iterable object consisting of a line per
    transaction
    """

    def split_records(self) -> Movimento:
        if self.excel_version == 1:
            return self._get_movimenti_V1()
        if self.excel_version == 2:
            return self._get_movimenti_V2()

    """
    Override method, use to generate Statement object
    """

    def parse(self):
        return super(IntesaSanPaoloXlsxParser, self).parse()

    """
    Override use to Parse given transaction line and return StatementLine
    object
    """

    def parse_record(self, mov: Movimento) -> StatementLine:
        logging.debug(mov.stat_line)
        return mov.stat_line

    def _get_account_id(self) -> str:
        if self.excel_version == 1:
            return self.wb['Lista Movimenti']['D8'].value
        elif self.excel_version == 2:
            # Remove the slash in order to make it consistent with the V1 version
            return self.wb['Lista Operazione']['C7'].value.split(" ")[1].replace('/', '')

    def _get_currency(self) -> str:
        # !!! Write All key value lower case !!!
        trans_map = {'euro': 'EUR',
                     'eur': 'EUR'}
        if self.excel_version == 1:
            val = self.wb['Lista Movimenti']['D22'].value
        elif self.excel_version == 2:
            # Suppose all operations have the same currency
            # Takes the currency from the first operation
            # this should be safe because there is always atleast 1 operation
            # since you can't export empty files
            val = self.wb['Lista Operazione']['G20'].value  # First operation
        return trans_map[val.lower()]

    def _get_start_balance(self) -> Decimal:
        if self.excel_version == 1:
            return Decimal(self.wb['Lista Movimenti']['E11'].value)
        elif self.excel_version == 2:
            return None

    def _get_end_balance(self) -> Decimal:
        if self.excel_version == 1:
            return Decimal(self.wb['Lista Movimenti']['E12'].value)
        elif self.excel_version == 2:
            return None

    def _get_start_date(self) -> datetime:
        if self.excel_version == 1:
            date = self.wb['Lista Movimenti']['D11'].value
            return datetime.strptime(date, '%d.%m.%Y')
        elif self.excel_version == 2:
            # On this version, C16 isn't always present, so calculate variation directly from record.
            # Operation are sort by date descending, so select last record
            try:
                date = datetime.strptime(self.wb['Lista Operazione']["C16"].value, '%d/%m/%Y')
            except (TypeError, ValueError):
                date = None
                # Find oldest date
                for row in self.wb['Lista Operazione'].iter_rows(20, values_only=True):
                    if row[0]:
                        date = row[0]
                    else:
                        break
            return date

    def _get_end_date(self) -> datetime:
        if self.excel_version == 1:
            date = self.wb['Lista Movimenti']['D12'].value
            return datetime.strptime(date, '%d.%m.%Y')
        elif self.excel_version == 2:
            # On this version, C17 isn't always present, so calculate
            # variation directly from record.
            # Operation are sort by date descending, so select first record
            try:
                date = datetime.strptime(self.wb['Lista Operazione']["C17"].value, '%d/%m/%Y')
            except (TypeError, ValueError):
                date = self.wb['Lista Operazione']["A20"].value
            return date

    """ Private method to parse all record lines """

    def _get_movimenti_V1(self) -> Iterator[Movimento_V1]:
        starting_column = column_index_from_string('A')
        ending_column = column_index_from_string('G')
        starting_row = 30

        for row in self.wb['Lista Movimenti']. \
                iter_rows(starting_row, None, starting_column, ending_column, values_only=True):
            if row[0]:
                logging.debug(row)
                yield Movimento_V1(*row)
            else:
                break

    def _get_movimenti_V2(self) -> Iterator[Movimento_V2]:
        starting_column = column_index_from_string('A')
        ending_column = column_index_from_string('H')
        starting_row = 20

        for row in self.wb['Lista Operazione']. \
                iter_rows(starting_row, None, starting_column, ending_column, values_only=True):
            if row[0]:
                # save only accounted transactions
                if row[4] == "NON CONTABILIZZATO":
                    continue
                else:
                    logging.debug(row)
                    yield Movimento_V2(*row)
            else:
                break
